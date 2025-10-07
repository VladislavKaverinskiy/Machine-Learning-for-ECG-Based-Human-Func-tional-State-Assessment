# -*- coding: utf8 -*-

import json
import os

import pandas as pd
import numpy as np
from sklearn.utils import shuffle
from sklearn.impute import KNNImputer
from sklearn.preprocessing import StandardScaler
from scipy import stats

from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def descriptive_stats(df, class_col='class', out_path='stats_report.xlsx', mapping_path='class_mapping.json'):
    """
    Calculates statistics (mean, std, median, minimum, maximum, skewness, kurtosis)
    for the entire dataset and for each class separately. Saves to separate Excel sheets.

      Parameters:
        df (pd.DataFrame): original DataFrame with classes and features (natural data).
        class_col (str): name of the column with classes.
        out_path (str): path to save the report (Excel).
        mapping_path (str): path to the JSON file mapping class numbers to their original names.
    """

    # Loading the class mapping
    if os.path.exists(mapping_path):
        with open(mapping_path, 'r', encoding='utf-8') as f:
            class_mapping = json.load(f)
        # Invert the dictionary: {1: "xls_base..."}
        class_names = {v: k for k, v in class_mapping.items()}
    else:
        print(f"Warning: Class mapping file '{mapping_path}' is not found.")
        class_names = {}

    # Divide into features and classes
    features = df.drop(columns=[class_col])
    classes = df[class_col].unique()

    def calc_stats(sub_df):
        desc = sub_df.describe().T[['mean', 'std', '50%', 'min', 'max']].rename(columns={'50%': 'median'})
        desc['skewness'] = sub_df.skew()
        desc['kurtosis'] = sub_df.kurtosis()
        return desc

    # Open ExcelWriter
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # General statistics
        overall_stats = calc_stats(features)
        overall_stats.reset_index(inplace=True)
        overall_stats.rename(columns={'index': 'feature'}, inplace=True)
        overall_stats.to_excel(writer, sheet_name='Overall', index=False)

        # Statistics for each class
        for cls in sorted(classes):
            df_cls = df[df[class_col] == cls].drop(columns=[class_col])
            stats_cls = calc_stats(df_cls)
            stats_cls.reset_index(inplace=True)
            stats_cls.rename(columns={'index': 'feature'}, inplace=True)

            sheet_name = class_names.get(cls, f'Class_{cls}')
            # Excel limitation: sheet names cannot exceed 31 characters.
            sheet_name = sheet_name[:31]

            stats_cls.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"The statistics report is saved in '{out_path}'")


def test_distributions(series):
    """
    Tests the series for three distributions: normal, lognormal, and uniform.
    Returns a dict with the p-value for each test.
    """

    data = series.dropna()
    if len(data) < 10:
        # Too little data for testing
        return {'normal': np.nan, 'lognormal': np.nan, 'uniform': np.nan}

    # 1. Normal distribution - Shapiro-Wilk test
    stat_norm, p_norm = stats.shapiro(data)

    # 2. Lognormal - check the logarithm of the data (should be positive)
    if (data <= 0).any():
        p_lognorm = np.nan
    else:
        log_data = np.log(data)
        stat_log, p_lognorm = stats.shapiro(log_data)

    # 3. Uniform distribution - Kolmogorov-Smirnov test with uniform distribution
    d_uniform, p_uniform = stats.kstest(data, 'uniform', args=(data.min(), data.max() - data.min()))

    return {'normal': p_norm, 'lognormal': p_lognorm, 'uniform': p_uniform}


def analyze_distributions(df, class_col='class'):
    """
    For each feature, it calculates the p-value of the tests and selects the best distribution.
    Returns a DataFrame with the results.
    """

    features = df.drop(columns=[class_col])
    results = []

    for col in features.columns:
        pvals = test_distributions(features[col])
        best_dist = max(pvals, key=lambda k: pvals[k] if not pd.isna(pvals[k]) else -1)
        results.append({
            'feature': col,
            'p_normal': pvals['normal'],
            'p_lognormal': pvals['lognormal'],
            'p_uniform': pvals['uniform'],
            'best_fit': best_dist
        })

    return pd.DataFrame(results)


def normalize_and_save_params(df, class_col='class', json_path='normalization_params.json'):
    """
    Normalizes the features in df using the standardization method (StandardScaler),
    saves the normalization parameters in JSON,
    returns a normalized DataFrame with the class column preserved.

      Parameters:
        df (pd.DataFrame): The original DataFrame with features and a class column.
        class_col (str): The name of the column with class labels.
        json_path (str): The path to the file for saving the normalization parameters.

      Returns:
        pd.DataFrame: The normalized DataFrame.
    """

    # Separating features from the class column
    features = df.drop(columns=[class_col])

    # Create a StandardScaler and train it
    scaler = StandardScaler()
    features_scaled = scaler.fit_transform(features)

    # Create a DataFrame with normalized features
    df_scaled = pd.DataFrame(features_scaled, columns=features.columns)

    # Adding the class column back
    df_scaled[class_col] = df[class_col].values

    # Save normalization parameters
    normalization_params = {
        'mean': dict(zip(features.columns, scaler.mean_)),
        'scale': dict(zip(features.columns, scaler.scale_))
    }
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(normalization_params, f, ensure_ascii=False, indent=4)

    return df_scaled


def is_valid_numeric_column(series, min_unique=10, max_nan_ratio=0.2):
    try:
        numeric_series = pd.to_numeric(series, errors='coerce')
        if numeric_series.isna().mean() > max_nan_ratio:
            return False
        if numeric_series.nunique(dropna=True) < min_unique:
            return False
        return True
    except:
        return False


def load_data_from_folder(folder_path):
    all_data = []
    first_valid_headers = None
    column_validity_per_file = []

    for file in os.listdir(folder_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file)
            df = pd.read_excel(file_path)

            # Delete the first 3 columns
            df = df.iloc[:, 3:]

            # Remember the column names from the first file
            if first_valid_headers is None:
                first_valid_headers = df.columns

            # Rename the columns for unification
            df.columns = [f'col_{i}' for i in range(df.shape[1])]

            # Convert all values ​​to numbers (errors in NaN)
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

            # Check the validity of each column in this file
            validity = {}
            for col in df.columns:
                validity[col] = is_valid_numeric_column(df[col])

            column_validity_per_file.append(validity)

            # Adding a class
            df['class'] = os.path.splitext(file)[0]

            all_data.append(df)

    # Determine the columns that are valid in all files
    # Get a list of all columns
    all_columns = set()
    for v in column_validity_per_file:
        all_columns.update(v.keys())
    all_columns = sorted(all_columns)

    # For each column, check that it is valid in each file
    valid_cols = []
    for col in all_columns:
        if all(col in validity and validity[col] for validity in column_validity_per_file):
            if col != 'class':  # don't exclude the class, it is added separately
                valid_cols.append(col)

    # Now merge the dataframes, leaving only valid columns + class
    filtered_dfs = []
    for df in all_data:
        filtered_dfs.append(df[valid_cols + ['class']])

    full_df = pd.concat(filtered_dfs, ignore_index=True)

    # Renaming columns correctly according to the first file
    if first_valid_headers is not None:
        header_mapping = {}
        for col in valid_cols:
            idx = int(col.split('_')[1])  # extract the number from 'col_number'
            if idx < len(first_valid_headers):
                header_mapping[col] = first_valid_headers[idx]
            else:
                header_mapping[col] = col  # If the index is out of range, leave it as is

        full_df.rename(columns=header_mapping, inplace=True)

    min_non_nan = int(0.2 * (full_df.shape[1] - 1))
    full_df = full_df.dropna(thresh=min_non_nan, axis=0).reset_index(drop=True)
    features = full_df.drop(columns=['class'])
    imputer = KNNImputer(n_neighbors=3)
    features_imputed = imputer.fit_transform(features)
    df_imputed = pd.DataFrame(features_imputed, columns=features.columns)
    df_imputed['class'] = full_df['class']

    # Shuffle
    df_imputed = shuffle(df_imputed, random_state=1024).reset_index(drop=True)

    return df_imputed


def drop_correlated_features(df, threshold=0.8):
    # Correlation matrix by absolute values
    corr_matrix = df.corr().abs()

    # The upper triangle of the matrix without the diagonal
    upper_tri = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))

    corr_pairs = list()
    for n, col_name in enumerate(upper_tri ):
        for m, val in enumerate(upper_tri[col_name]):
            if val > threshold:
                corr_pairs.append([col_name, upper_tri.columns[m], val])
    print(corr_pairs)
    with open('strong_corr_pairs.txt', 'w', encoding='utf-8') as f:
        json.dump(corr_pairs, f, ensure_ascii=False, indent=4)

    upper_tri.to_excel('corr_upper_tri.xlsx', index=False, engine='openpyxl')

    # List of features to remove
    to_drop = [column for column in upper_tri.columns if any(upper_tri[column] > threshold)]

    # Removing highly correlated features
    df_reduced = df.drop(columns=to_drop)

    return df_reduced, to_drop


def normalize_vector(vec: dict) -> dict:
    """
    vec — dictionary {feature: value}
    Returns a normalized dictionary.
    """
    with open('normalization_params.json', 'r', encoding='utf-8') as f:
        params = json.load(f)

    means = params['mean']
    scales = params['scale']
    norm_vec = {}
    for feature, value in vec.items():
        mean = means.get(feature, 0)
        scale = scales.get(feature, 1)
        norm_vec[feature] = (value - mean) / scale if scale != 0 else 0
    return norm_vec


def cap_outliers_iqr(df):
    features = df.drop(columns=['class'])
    labels = df['class']
    for col in features.columns:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR

        df[col] = df[col].clip(lower=lower_bound, upper=upper_bound)
    df['class'] = labels.values
    return df


def drop_constant_columns(df, class_col='class'):
    # Exclude column with classes
    features = df.drop(columns=[class_col])
    # Identify columns that have more than 1 unique value
    non_constant_cols = [col for col in features.columns if features[col].nunique(dropna=False) > 1]
    # Create a dataframe with these columns + the class column
    df_filtered = df[non_constant_cols + [class_col]].copy()
    return df_filtered


def apply_lognormal_transform(df: pd.DataFrame, dist_results: pd.DataFrame, dist_column='best_fit',
                              feature_column='feature') -> pd.DataFrame:
    """
    Replaces log-normally distributed feature values ​​with their logarithms.

        Parameters:
        - df: pd.DataFrame — original data (before normalization).
        - dist_results: pd.DataFrame — table with distribution fitting results.
        - dist_column: str — name of the column with distribution types (default: 'best_fit').
        - feature_column: str — name of the column with feature names (default: 'feature').

        Returns:
        - pd.DataFrame — copy of df with log-transformed features.
    """
    df_transformed = df.copy()
    lognormal_features = dist_results[dist_results[dist_column] == 'lognormal'][feature_column].tolist()

    for feature in lognormal_features:
        if feature not in df_transformed.columns:
            continue  # If the column is not in the data, skip it
        series = df_transformed[feature]
        if (series <= 0).any():
            shift = abs(series.min()) + 1e-6
            df_transformed[feature] = np.log(series + shift)
        else:
            df_transformed[feature] = np.log(series + 1e-6)

    return df_transformed


def compare_class_means(df, class_col='class', class_mapping_path='class_mapping.json',
                        out_path='class_mean_comparison.xlsx'):
    """
    Compares the mean feature values ​​between all pairs of classes.
    Uses the Welch's t-test to test significance.
    Saves the results in Excel, with a separate sheet for each feature.
    Significant and marginal differences are highlighted.

    Parameters:
        df (pd.DataFrame): data (before normalization).
        class_col (str): name of the column with class labels.
        class_mapping_path (str): path to the JSON file with the class mapping.
        out_path (str): path to the Excel file.
    """
    # Loading the class display
    if os.path.exists(class_mapping_path):
        with open(class_mapping_path, 'r', encoding='utf-8') as f:
            mapping = json.load(f)
        # Invert it for convenience: class number -> name
        class_names = {v: k for k, v in mapping.items()}
    else:
        print(f"Class mapping file '{class_mapping_path}' is not found. Using class numbers.")
        unique_classes = df[class_col].unique()
        class_names = {cls: f'class_{cls}' for cls in unique_classes}

    features = df.drop(columns=[class_col]).columns
    classes = sorted(df[class_col].unique())

    wb = Workbook()
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # p < 0.05
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 0.05 <= p < 0.1

    # If there is a default sheet in the book, delete it.
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    for feature in features:
        rows = []
        # Table header
        rows.append(
            ['Class 1', 'Class 2', f'Mean {feature} Class 1', f'Mean {feature} Class 2', 't-statistic', 'p-value'])

        for cls1, cls2 in combinations(classes, 2):
            data1 = df.loc[df[class_col] == cls1, feature].dropna()
            data2 = df.loc[df[class_col] == cls2, feature].dropna()

            mean1 = data1.mean() if len(data1) > 0 else np.nan
            mean2 = data2.mean() if len(data2) > 0 else np.nan

            if len(data1) < 3 or len(data2) < 3:
                t_stat = np.nan
                p_val = np.nan
            else:
                t_stat, p_val = stats.ttest_ind(data1, data2, equal_var=False)

            rows.append([
                class_names.get(cls1, cls1),
                class_names.get(cls2, cls2),
                round(mean1, 5) if not pd.isna(mean1) else '',
                round(mean2, 5) if not pd.isna(mean2) else '',
                round(t_stat, 5) if not pd.isna(t_stat) else '',
                round(p_val, 5) if not pd.isna(p_val) else ''
            ])

        ws = wb.create_sheet(title=feature[:31])

        # Writing data to a sheet
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Apply shading to the p-value (column 6), starting from the 2nd row
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                try:
                    p = float(cell.value)
                    if p < 0.05:
                        cell.fill = red_fill
                    elif p < 0.1:
                        cell.fill = yellow_fill
                except:
                    pass

        # Expand the column width a little for readability.
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(out_path)
    print(f"Comparison of class averages is saved in '{out_path}'")


if __name__ == "__main__":
    data_folder_path = 'data'
    out_data_folder_path = 'work_data'
    final_df = load_data_from_folder(data_folder_path)

    df_capped = cap_outliers_iqr(final_df.copy())

    df_features = df_capped.drop(columns=['class'])
    df_cleaned, dropped = drop_correlated_features(df_features, threshold=0.8)

    print(f"Features removed due to strong correlation: {dropped}")
    with open('dropped_features.txt', 'w', encoding='utf-8') as f:
        json.dump([dropped], f, ensure_ascii=False, indent=4)

    df_cleaned['class'] = df_capped['class']

    print(df_cleaned.head())
    unique_labels = sorted(df_cleaned['class'].unique())
    label_to_number = {label: idx + 1 for idx, label in enumerate(unique_labels)}
    with open('class_mapping.json', 'w', encoding='utf-8') as f:
        json.dump(label_to_number, f, ensure_ascii=False, indent=4)
    df_cleaned['class'] = df_cleaned['class'].map(label_to_number)

    df_cleaned = drop_constant_columns(df_cleaned)

    descriptive_stats(df=df_cleaned, class_col='class', out_path='stats_report.xlsx', mapping_path='class_mapping.json')

    dist_results = analyze_distributions(df_cleaned)
    print(dist_results)
    # dist_results.to_excel('distribution_analysis.xlsx', index=False, engine='openpyxl')

    final_df_transformed = apply_lognormal_transform(df_cleaned, dist_results)

    # compare_class_means(df=df_cleaned, class_col='class', out_path='class_mean_comparison.xlsx')
    # final_df_transformed.to_csv(out_data_folder_path + '/preprocessed.csv', index=False, encoding='utf-8')
    # final_df_transformed.to_excel(out_data_folder_path + '/preprocessed.xlsx', index=False, engine='openpyxl')

    df_normalized = normalize_and_save_params(final_df_transformed)

    # df_normalized.to_csv(out_data_folder_path + '/normalized..csv', index=False, encoding='utf-8')
    # df_normalized.to_excel(out_data_folder_path + '/normalized..xlsx', index=False, engine='openpyxl')
