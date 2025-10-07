# -*- coding: utf8 -*-

import json
import os
import copy
import re

import pandas as pd
import numpy as np
from sklearn.utils import shuffle
from sklearn.impute import KNNImputer
from sklearn.preprocessing import StandardScaler
from scipy import stats

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


def descriptive_stats(df_list, class_col='class', out_path='stats_report_clusters.xlsx', class_mapping_rev=None):
    def calc_stats(sub_df):
        desc = sub_df.describe().T[['mean', 'std', '50%', 'min', 'max']].rename(columns={'50%': 'median'})
        desc['skewness'] = sub_df.skew()
        desc['kurtosis'] = sub_df.kurtosis()
        return desc

    def clean_sheet_name(name):
        return re.sub(r'[\\/*?:\[\]]', '_', name)[:31]

    # Collect all the data with the cluster into one table to calculate the distribution
    combined_list = []
    for cluster_name, df in df_list.items():
        tmp = df.copy()
        tmp['cluster'] = cluster_name  # Adding a column with a cluster
        combined_list.append(tmp)
    combined_df = pd.concat(combined_list, ignore_index=True)

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for cluster_name, df in df_list.items():
            features = df.drop(columns=[class_col])

            # Statistics of features in a specific cluster
            stats_cls = calc_stats(features)
            stats_cls.reset_index(inplace=True)
            stats_cls.rename(columns={'index': 'feature'}, inplace=True)

            # Distribution of classes within a given cluster (percentage of cluster)
            class_counts = df[class_col].value_counts(normalize=True) * 100
            if class_mapping_rev:
                class_distribution = {
                    class_mapping_rev.get(int(cls), f'class_{cls}'): round(pct, 2)
                    for cls, pct in class_counts.items()
                }
            else:
                class_distribution = {
                    f'class_{int(cls)}': round(pct, 2)
                    for cls, pct in class_counts.items()
                }

            distribution_row = {'feature': 'class_distribution_in_cluster'}
            distribution_row.update(class_distribution)
            stats_cls = pd.concat([stats_cls, pd.DataFrame([distribution_row])], ignore_index=True)

            sheet_name = clean_sheet_name(f'Cluster_{cluster_name}')
            stats_cls.to_excel(writer, sheet_name=sheet_name, index=False)

        # Calculate the inverse distribution:
        # for each class, what percentage of its objects fell into each cluster
        cross = pd.crosstab(combined_df[class_col], combined_df['cluster'])
        cross_perc = cross.div(cross.sum(axis=1), axis=0) * 100

        if class_mapping_rev:
            cross_perc.index = cross_perc.index.map(lambda x: class_mapping_rev.get(int(x), f'class_{x}'))

        cross_perc = cross_perc.round(2)

        # Let's write it down on a separate sheet
        matrix_sheet = clean_sheet_name('Class_to_Cluster_Distribution')
        cross_perc.to_excel(writer, sheet_name=matrix_sheet)

    print(f"The statistics report is saved in '{out_path}'")


def compare_clusters_means(df_dict, class_col='class', out_path='mean_comparison.xlsx'):
    cluster_names = list(df_dict.keys())
    all_features = set()

    for df in df_dict.values():
        all_features.update(df.select_dtypes(include='number').columns.difference([class_col]))

    all_features = sorted(all_features)

    wb = Workbook()
    # Delete the default sheet (so there is no empty 'Sheet')
    default_sheet = wb.active
    wb.remove(default_sheet)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # p < 0.01
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # p < 0.05

    for feat in all_features:
        means = {}
        data_groups = []
        for name in cluster_names:
            df = df_dict[name]
            if feat in df.columns:
                values = df[feat].dropna().values
                if len(values) > 1:
                    means[name] = np.mean(values)
                    data_groups.append(values)
                else:
                    means[name] = np.nan
            else:
                means[name] = np.nan

        # ANOVA
        if len(data_groups) >= 2:
            _, p_anova = stats.f_oneway(*data_groups)
            anova_p = round(p_anova, 5)
        else:
            anova_p = 'insufficient'

        # Form a list of lines for the sheet (two columns: parameter and value)
        rows = []

        # Cluster averages
        rows.append(('Average values ​​by cluster:', ''))
        for name in cluster_names:
            rows.append((f'  {name}', means.get(name, 'missing')))

        # ANOVA p-value
        rows.append(('ANOVA p-value', anova_p))

        # Pairwise t-tests and flags
        rows.append(('Pairwise t-tests (p-value) and flags:', ''))
        for i in range(len(cluster_names)):
            for j in range(i + 1, len(cluster_names)):
                c1, c2 = cluster_names[i], cluster_names[j]
                pval_col = f'{c1} vs {c2} p-value'
                large_diff_col = f'{c1} vs {c2} large diff (>=1.5)'
                opp_sign_col = f'{c1} vs {c2} opposite sign'

                if feat in df_dict[c1].columns and feat in df_dict[c2].columns:
                    data1 = df_dict[c1][feat].dropna()
                    data2 = df_dict[c2][feat].dropna()

                    if len(data1) > 1 and len(data2) > 1:
                        _, pval = stats.ttest_ind(data1, data2, equal_var=False)
                        pval_rnd = round(pval, 5)

                        mean1 = np.mean(data1)
                        mean2 = np.mean(data2)

                        # Large difference
                        if (mean1 == 0 and abs(mean2) >= 1.5) or (mean2 == 0 and abs(mean1) >= 1.5):
                            large_diff = True
                        elif mean1 != 0 and mean2 != 0:
                            ratio = abs(mean1 / mean2)
                            large_diff = ratio >= 1.5 or (1/ratio) >= 1.5
                        else:
                            large_diff = False

                        # Opposite sign
                        opp_sign = (mean1 * mean2) < 0

                        rows.append((pval_col, pval_rnd))
                        rows.append((large_diff_col, large_diff))
                        rows.append((opp_sign_col, opp_sign))
                    else:
                        rows.append((pval_col, 'insufficient'))
                        rows.append((large_diff_col, 'insufficient'))
                        rows.append((opp_sign_col, 'insufficient'))
                else:
                    rows.append((pval_col, 'missing'))
                    rows.append((large_diff_col, 'missing'))
                    rows.append((opp_sign_col, 'missing'))

        # Create a sheet
        ws = wb.create_sheet(title=feat[:31])
        for row in rows:
            ws.append(row)

        # Stylizing p-values
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                try:
                    val = float(cell.value)
                    # Color only the p-value columns (they are in the second column,
                    # and the first row has 'p-value' in the title)
                    key_cell = ws.cell(row=cell.row, column=1)
                    if 'p-value' in str(key_cell.value):
                        if val < 0.01:
                            cell.fill = red_fill
                        elif val < 0.05:
                            cell.fill = yellow_fill
                except:
                    pass

    wb.save(out_path)
    print(f"The results of the comparison of averages are recorded in '{out_path}'")


def test_distributions(series):
    """
    Tests a series for normality, lognormality, and uniformity.
    Returns a dictionary with the p-value and the best-fitting distribution.
    """
    data = series.dropna()
    if len(data) < 10:
        return {'normal': np.nan, 'lognormal': np.nan, 'uniform': np.nan, 'best_fit': None}

    # Normal distribution
    if data.min() == data.max():
        p_norm = np.nan
    else:
        try:
            stat_norm, p_norm = stats.shapiro(data)
        except Exception:
            p_norm = np.nan

    # Lognormal
    if (data <= 0).any() or data.min() == data.max():
        p_lognorm = np.nan
    else:
        log_data = np.log(data)
        if log_data.min() == log_data.max():
            p_lognorm = np.nan
        else:
            try:
                stat_log, p_lognorm = stats.shapiro(log_data)
            except Exception:
                p_lognorm = np.nan

    # Uniform
    if data.min() == data.max():
        p_uniform = np.nan
    else:
        try:
            d_uniform, p_uniform = stats.kstest(data, 'uniform', args=(data.min(), data.max() - data.min()))
        except Exception:
            p_uniform = np.nan

    # Selecting the best distribution by maximizing p-value
    p_values = {
        'normal': p_norm,
        'lognormal': p_lognorm,
        'uniform': p_uniform
    }
    best_fit = max(p_values.items(), key=lambda x: (x[1] if not np.isnan(x[1]) else -1))[0] if any(
        not np.isnan(v) for v in p_values.values()) else None

    return {**p_values, 'best_fit': best_fit}


def generate_distribution_report(df_dict, output_path="clusters_distribution_tests.xlsx"):
    cluster_names = [os.path.splitext(os.path.basename(k))[0] for k in df_dict.keys()]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for name, df in zip(cluster_names, df_dict.values()):
            result = {}
            for col in df.select_dtypes(include=[np.number]).columns:
                result[col] = test_distributions(df[col])
            result_df = pd.DataFrame(result).T  # Транспонируем: строки — признаки, столбцы — типы распределений
            result_df.to_excel(writer, sheet_name=name[:31])  # Excel ограничивает имя листа до 31 символа

    print(f"Distributions are saved in {output_path}")


def find_lognormal_features_across_clusters(df_dict, class_col='class'):
    lognormal_features = set()
    for df in df_dict.values():
        features = df.select_dtypes(include='number').columns.difference([class_col])
        for feature in features:
            dist_res = test_distributions(df[feature])
            if dist_res['best_fit'] == 'lognormal':
                lognormal_features.add(feature)
    return lognormal_features


def preprocess_lognormal_features_consistent(df_dict, lognormal_features, class_col='class'):
    processed = {}
    for cluster_name, df in df_dict.items():
        df_copy = df.copy()
        for feature in lognormal_features:
            if feature in df_copy.columns:
                series = df_copy[feature]
                if series.min() <= 0:
                    shift = abs(series.min()) + 1e-6
                    df_copy[feature] = np.log(series + shift)
                else:
                    df_copy[feature] = np.log(series)
        processed[cluster_name] = df_copy
    return processed




def load_data(folder_path):
    clusters_data = dict()
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file)
            df = pd.read_excel(file_path)
            clusters_data[os.path.splitext(os.path.basename(file_path))[0]] = copy.deepcopy(df)
    return clusters_data



if __name__ == "__main__":
    data_folder_path = 'clusters_output'
    case = "4 clusters"
    input_clusters = load_data(data_folder_path + "/" + case)

    with open("class_mapping.json", "r", encoding="utf-8") as f:
        class_mapping = json.load(f)
    class_mapping_rev = {v: k for k, v in class_mapping.items()}

    descriptive_stats(input_clusters, class_col='class', out_path='stats_report_clusters.xlsx',
                      class_mapping_rev=class_mapping_rev)

    generate_distribution_report(input_clusters)


    lognormal_feats = find_lognormal_features_across_clusters(input_clusters, class_col='class')
    transformed_clusters = preprocess_lognormal_features_consistent(input_clusters, lognormal_feats, class_col='class')

    compare_clusters_means(transformed_clusters, out_path='clusters_mean_comparison.xlsx')
